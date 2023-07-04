#!/usr/bin/python
import os
import csv
import time
import threading
import openpyxl
import concurrent.futures
from subprocess import run
from pprint import pprint
from androguard import misc

"""
### Variable Declaration
"""

# dirPath = os.getcwd() + "2023_REU_Workspace/covid_19_Dataset" # For when I wanna run in terminal
dirPath = "2023_REU_Workspace/covid_19_Dataset" # Codium executive path
directory = os.listdir(dirPath) # Get's child directories to iterate through, I really wish to consolidate this further
keys = ["Application Name", "Package Name", "APK File", "AV Rank", "Cloned", "Apps sharing APK", "Total permission requests"] # Reference keys for apk data
dummy = { # To demonstrate how I wanted to structure our data for the csv files
    "App Name" : {
        "256hash.apk": {
            "pkg name": "name of apk's package",
            "permissions": ["Permission spread for the given apk file"],
            "avRank": "integer > -1 representing AV Rank collected during COVIDMalware.pdf study",
            "cloned": "boolean determining if this apk has been cloned",
            "clones": "array of strings of apps that also use this apk file"
        }
    }
}

permSpread = [] # list of all permissions requested across all APK files.
files = [] # declaring list for apk
apps = {} # Declaring dictionary for APK data

"""
### Function Declaration
"""

def unzip(filePath):
    print(f"# CONSOLE: Unzipping {filePath}...")

    try:
        run(["unzip", filePath, "-d", dirPath])
        print(f"# CONSOLE: unzipped {filePath}. Deleting zipfiles...")
        os.remove(filePath)
    except:
        print(f"# CONSOLE: Unable to unzip {filePath}.")

def unzipDir(dirPath):
    zips = []

    # Filters out Zip files within directory
    for file in directory:
        if len(os.path.splitext(file)[1]) != 0:
            filePath = os.path.join(dirPath, file)
            zips.append(filePath) if os.path.splitext(file)[1] == ".zip" else print(f"{filePath} is not a zip file")
    
    t0 = time.time()

    ### DANGER: MULTITHREADING
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(unzip, zips)

    print(f"Analysis Duration: {(time.time()-t0):.2f} second(s)") # Time the performace of my program

def apkAnalyzer(apk): # Calls Androguard's misc.AnalyzeAPK() and updates declared dictionary
    print(f"### CONSOLE: Analyzing file: {apk}...")
    global apps

    # Attempt to call Androguard's APK Scanner
    try:
        a, b, c = misc.AnalyzeAPK(apk)
            
        name = a.get_app_name()
        pkg = a.get_package()
        permissions = a.get_permissions()
        apk = apk.split("/")[-1]
        avRank = 0
        cloned = 0
        clones = []

        # Checking if duplicate APKs are found and which apps use them
        for app in apps:
            if apk in apps[app].keys():
                # For the current APK
                cloned = 1 
                clones.append(app)

                # if other APKs are found, update them too
                apps[app][apk]["clones"].append(name)
                apps[app][apk]["cloned"] = 1

        # Collect AV Rank from COVIDMalware.pdf data
        wb = openpyxl.load_workbook(os.path.join(dirPath, "covid19apps.xlsx"))
        for sheet in wb.worksheets:
            for r in range(2, sheet.max_row):
                if sheet.cell(r, 6).value in apk:
                    avRank = sheet.cell(r, 5).value

        # Update Apps dictionary
        if name not in apps:
            apps[name] = {
                apk: {
                    "pkg name": pkg,
                    "permissions": permissions,
                    "avRank": avRank,
                    "cloned": cloned,
                    "clones": clones,
                }
            }
        else:
            if apk not in apps[name]:
                apps[name][apk] = {
                    "pkg name": pkg,
                    "permissions": permissions,
                    "avRank": avRank,
                    "cloned": cloned,
                    "clones": clones,
                }
        print(f"### CONSOLE: Completed analysis.")
    except Exception as ex:
        print(f"### CONSOLE: Error while analyzing {apk}. Aborting attempt.")
        # Seeing how many cause us problems and why
        with open(os.path.join(os.getcwd(), "AnalysisFailures.txt"),"a+") as outFile:
            outFile.write(f"{apk} failed with error: {ex}\n")

"""
### INIT MAIN
"""
if __name__ == '__main__':
    """
    ### Compiles list of apk files
    """
    # Really wish I would've condensed both of these loops of 'directory' into a single block
    for file in directory:
        if ".zip" in os.path.splitext(file)[1]:
            print("### CONSOLE: ZipFiles Found.")
            unzipDir(dirPath) # WARNING: THIS WILL UNZIP ALL FILES WITHIN THE DIRECTORY

    # Bonus points for reducing time compelxity here too
    for item in directory:
        if len(os.path.splitext(item)[1]) == 0:
            for file in os.listdir(os.path.join(dirPath, item)):
                files.append(os.path.join(dirPath, item, file))

    with open(os.path.join(os.getcwd(),"APKsFound.txt"), "w") as outFile:
        for file in files:
            outFile.write(file+"\n")

    """
    ### APK Analysis
    """
    t0 = time.time() # Start timer

    ### DEBUG: Single file analysis
    # apkAnalyzer("2023_REU_Workspace/covid_19_Dataset/covid19apps_0526_0/40d02cdcf32b40eac293058b46c24c35b8702f7b6f91e1176bdda0ede93c29cc.apk")
    
    # iterative execution
    # for file in files:
    #     apkAnalyzer(file)
    # Multithreaded execution
    with concurrent.futures.ThreadPoolExecutor() as executor: # No restrictions with how much threads this program can run
        executor.map(apkAnalyzer, files) # Analyzing entirety of files array

    print(f"Analysis Duration: {(time.time()-t0):.2f} second(s)") # Time the performace of my program

    """
    ### Updating keys array
        - to include all permissions found during our APK analysis
    """
    # ### DEBUG: Pre-structured Apps dictionary
    # apps = {}

    # pprint(apps)

    # Aggregating all requested permissions
    for app in apps:
        for key, values in apps[app].items():
            for i in values["permissions"]:
                permSpread.append(i) if i not in permSpread else print("",end="")

            
    # print(permSpread)
    permSpread.sort()
    keys += permSpread
    # print(keys[7:])
    
    """
    ### Writing to CSV
    """
    with open(os.path.join(os.getcwd(), "COVID19_APK_Data_06-2023.csv"), 'w') as outFile:
        writer = csv.writer(outFile)
        writer.writerow(keys)
        for app in apps:
            for apk, values in apps[app].items():
                line = [] # THIS FUCK WAS IN THE OUTER LOOP AND DUPLICATED MY CSV LINES
                # print(f"Application Name: {app}")
                line.append(app)
                # print("Package Name: ", values["pkg name"])
                line.append(values["pkg name"])
                # print(f"APK File: {key}")
                line.append(apk)
                # print("AV Rank: ", values["avRank"])
                line.append(values["avRank"])
                # print(f"{apk} has been cloned.") if values["cloned"] > 0 else print("", end="")
                line.append(values["cloned"])
                # print(f"Apps sharing {apk}: ", values["clones"])
                line.append(values["clones"])
                # print("Total permssion requests: ", len(values["permissions"]))
                line.append(len(values["permissions"]))
                # print("Permissions spread: ", values["permissions"])
                for i in keys[7:]:
                    line.append(1) if i in values["permissions"] else line.append(0)

                # Will not include the Malware (y/n) col, since I feel the AV Rank satisfies that metric
                # print(f"\n###{key} is malware!!###\n") if values["avRank"] > 0 else print("APK is benign.")
                
                # print(line)
                writer.writerow(line)
    
    # print("Current datastructure Format:")
    # pprint(dummy)

    """
    ### Problems and errors
    - My spreadsheet's all screwed up, I coded it wrong :(
        -   I KNOW WHY: THE ARRAY DECLARATION WAS IN ONE LOOP TO HIGH:
    - Androguard analyzeAPK() throws error on 40d02cdcf32b40eac293058b46c24c35b8702f7b6f91e1176bdda0ede93c29cc.apk in covid_19_526_0 
        - "Exception has occurred: BadZipFile File is not a zip file" --> Show results.txt to verify apk isn't corrupted
        - idfk why I'm getting that error
    """

    """
    ### Notes on what to do
    - Machine Learning part --> 90% accuracy
        - compare N-B, SVM, DTrees
        - metrics comparison: f1, percision, accuracy, recall, time, etc.
        - K-Fold Cross-Validation to partition and work through our full dataset
        
    - NN part --> Shoot for 98% accuracy
        - passing to DNN
        - converting to image map for C/ANN
        
    - For the best model
        - Fine tune it!!!
        
    - Step 3
        - SMOTE to assist balancing dataset
            - 5 with SMOTE applied, 5 without SMOTE applied
        - Comparing to prepared xlsx given
        
    - replicate bar graph from COVIDMalware.pdf representing how many apps use a specific permission

    - graph how many apps use a quantity of permissions
        - sum all 1 values in permissions in csv 
    """