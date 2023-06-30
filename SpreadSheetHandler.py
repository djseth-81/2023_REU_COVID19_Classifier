#!/usr/bin/python
import os
import openpyxl
import csv
from pprint import pprint

wb = openpyxl.load_workbook("2023_REU_Workspace/COVID_malware_permissions_all.xlsx")

apps = {}

dummy = {
        "App Name" : {
            "256hash.apk": {
                "pkg name": "name of apk's package",
                "permissions": ["Permission spread for the given apk file"],
                "avRank": "integer > -1 representing AV Rank collected during COVIDMalware.pdf study",
                "cloned": "boolean determining if this apk has been cloned",
                "clones": "array of strings of apps that also use this apk file"
            }
        }
    }
# print("Current datastructure Format:")
# pprint(dummy)

for sheet in wb.worksheets:
    permSpread = [sheet.cell(1, col).value for col in range(3, sheet.max_column)]
    # print(permSpread)

    for r in range(2, 3):
        apps[sheet.cell(r, 1).value] = {
            "malware": sheet.cell(r, 2).value,
            # "Permissions": [sheet.cell(r, c).value for c in range(3, sheet.max_column)]
            "Permissions": [permSpread[i] if sheet.cell(r, i).value > 0 else "fuck" for i in range(3, sheet.max_column)]
        }

for item in apps.values():
    print(len(item["Permissions"]))
    for i in range(len(item["Permissions"]) - 1):
        print(item["Permissions"][i])
        if "fuck" == item["Permissions"][i]:
            item["Permissions"].pop(i)

    print(item["Permissions"])

# pprint(apps)